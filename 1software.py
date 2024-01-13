from tkinter import*
import math,random
from tkinter import messagebox
import os
from twilio.rest import Client
import openpyxl ,xlrd
from openpyxl import Workbook, load_workbook
import pathlib
from datetime import date

class Bill_App:
    def __init__(self,root):
        self.root=root
        self.root.geometry("1350x800+0+0")
        self.root.title("Billing software")
        bg_color="#074466"
        title=Label(self.root,text="Billing Software",bd=12,relief=GROOVE,bg=bg_color,fg="white",font=("times new roman",30,"bold"),pady=2).pack(fill=X)
        
        #variables

        #buffalo
        self.bweight=DoubleVar()
        self.bfat=DoubleVar()
        self.btemp=IntVar()
        self.bsnf=IntVar()
        self.bwater=IntVar()
        self.brate=IntVar()
        self.baverage=IntVar()

        #cow
        self.cweight=DoubleVar()
        self.cfat=DoubleVar()
        self.ctemp=IntVar()
        self.csnf=IntVar()
        self.cwater=IntVar()
        self.crate=IntVar()
        self.caverage=IntVar()

        #total
        self.baffelo_price=StringVar()
        self.cow_price=StringVar()
        self.totalboth=StringVar()
        self.baffelo_fat=StringVar()
        self.cow_fat=StringVar()
        self.total_bill=StringVar()
        
        
        #customer
        self.c_name=StringVar()
        self.c_phon=StringVar()
        self.bill_no=IntVar()
        self.search_bill=StringVar()
        self.Date=StringVar()
        self.kal=StringVar()
        self.copy=StringVar()
        self.copy1=StringVar()
        self.copy2=StringVar()
        self.copy3=StringVar()
         
        today = date.today()
        d1 = today.strftime("%d/%m/%y")
        self.Date.set(d1)

        #customer details
        F1=LabelFrame(self.root,bd=15,relief=GROOVE,text="Customer details",font=("times new roman",15,"bold"),fg="gold",bg=bg_color)
        F1.place(x=0,y=75,relwidth=1)

        cname_lb1=Label(F1,text="Customer Name",bg=bg_color,fg="white",font=("times new roman",18,"bold")).grid(row=0,column=0,padx=20,pady=5)
        cname_txt=Entry(F1,width=15,textvariable=self.c_name,font="arial 15",bd=7,relief=SUNKEN).grid(row=0,column=1,pady=5,padx=10)

        cphn_lb1=Label(F1,text="Phone No.",bg=bg_color,fg="white",font=("times new roman",18,"bold")).grid(row=0,column=2,padx=20,pady=5)
        cphn_txt=Entry(F1,width=15,textvariable=self.c_phon,font="arial 15",bd=7,relief=SUNKEN).grid(row=0,column=3,pady=5,padx=10)

        c_bill_lb1=Label(F1,text="Bill No.",bg=bg_color,fg="white",font=("times new roman",18,"bold")).grid(row=0,column=4,padx=20,pady=5)
        c_bill_txt=Entry(F1,width=15,textvariable=self.search_bill,font="arial 15",bd=7,relief=SUNKEN).grid(row=0,column=5,pady=5,padx=10)

        bill_btn=Button(F1,text="search",command=self.fetch,width=15,bd=7,font="arial 12 bold").grid(row=0,column=6,padx=40,pady=10)
        self.bill_n()
        #buffalo
        F2=LabelFrame(self.root,bd=10,relief=GROOVE,text="Baffelo ",font=("times new roman",15,"bold"),fg="gold",bg=bg_color)
        F2.place(x=5,y=180,width=400,height=365)

        per_lbl=Label(F2,text="weight",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=0,column=0,padx=5,pady=5,sticky="w")
        per_text=Entry(F2,width=10,textvariable=self.bweight,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=0,column=1,padx=5,pady=5)
        
        fat_lbl=Label(F2,text="Fat",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=1,column=0,padx=5,pady=5,sticky="w")
        fat_text=Entry(F2,width=10,textvariable=self.bfat,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=1,column=1,padx=5,pady=5)

        temp_lbl=Label(F2,text="Temp",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=2,column=0,padx=5,pady=5,sticky="w")
        temp_text=Entry(F2,width=10,textvariable=self.btemp,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=2,column=1,padx=5,pady=5)

        snf_lbl=Label(F2,text="S N F",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=3,column=0,padx=5,pady=5,sticky="w")
        snf_text=Entry(F2,width=10,textvariable=self.bsnf,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=3,column=1,padx=5,pady=5)

        water_lbl=Label(F2,text="Water",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=4,column=0,padx=5,pady=5,sticky="w")
        water_text=Entry(F2,width=10,textvariable=self.bwater,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=4,column=1,padx=5,pady=5)

        rate_lbl=Label(F2,text="Rate",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=5,column=0,padx=5,pady=5,sticky="w")
        rate_text=Entry(F2,width=10,textvariable=self.brate,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=5,column=1,padx=5,pady=5)

        ave_lbl=Label(F2,text="Average",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=6,column=0,padx=5,pady=5,sticky="w")
        ave_text=Entry(F2,width=10,textvariable=self.baverage,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=6,column=1,padx=5,pady=5)

        #cow
        F3=LabelFrame(self.root,bd=10,relief=GROOVE,text="Cow ",font=("times new roman",15,"bold"),fg="gold",bg=bg_color)
        F3.place(x=410,y=180,width=400 ,height=365)

        per_lbl=Label(F3,text="weight",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=0,column=0,padx=5,pady=5,sticky="w")
        per_text=Entry(F3,width=10,textvariable=self.cweight,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=0,column=1,padx=5,pady=5)
        
        fat_lbl=Label(F3,text="Fat",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=1,column=0,padx=5,pady=5,sticky="w")
        fat_text=Entry(F3,width=10,textvariable=self.cfat,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=1,column=1,padx=5,pady=5)

        temp_lbl=Label(F3,text="Temp",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=2,column=0,padx=5,pady=5,sticky="w")
        temp_text=Entry(F3,width=10,textvariable=self.ctemp,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=2,column=1,padx=5,pady=5)

        snf_lbl=Label(F3,text="S N F",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=3,column=0,padx=5,pady=5,sticky="w")
        snf_text=Entry(F3,width=10,textvariable=self.csnf,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=3,column=1,padx=5,pady=5)

        water_lbl=Label(F3,text="Water",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=4,column=0,padx=5,pady=5,sticky="w")
        water_text=Entry(F3,width=10,textvariable=self.cwater,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=4,column=1,padx=5,pady=5)

        rate_lbl=Label(F3,text="Rate",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=5,column=0,padx=5,pady=5,sticky="w")
        rate_text=Entry(F3,width=10,textvariable=self.crate,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=5,column=1,padx=5,pady=5)

        ave_lbl=Label(F3,text="Average",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=6,column=0,padx=5,pady=5,sticky="w")
        ave_text=Entry(F3,width=10,textvariable=self.caverage,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=6,column=1,padx=5,pady=5)

        #bill area-
        F4=Frame(self.root,bd=10,relief=GROOVE)
        F4.place(x=890,y=180,width=400 ,height=365)
        bill_title=Label(F4,text="Bill Area",font="arial 15 bold",bd=7,relief=GROOVE).pack(fill=X)
        scrol_y=Scrollbar(F4,orient=VERTICAL)
        self.txtarea=Text(F4,yscrollcommand=scrol_y.set)
        scrol_y.pack(side=RIGHT,fill=Y)
        scrol_y.config(command=self.txtarea.yview)
        self.txtarea.pack(fill=BOTH,expand=1)

        #button frame
        F5=LabelFrame(self.root,bd=10,relief=GROOVE,text="Bill Menu",font=("times new roman",15,"bold"),fg="gold",bg=bg_color)
        F5.place(x=0,y=550,relwidth=1,height=150)

        m1_lbl=Label(F5,text="Total Bufflo Milk Price",bg=bg_color,fg="White",font=("times new roman",13,"bold")).grid(row=0,column=0,padx=4,pady=4,sticky="w")
        m1_txt=Entry(F5,width=20,textvariable=self.baffelo_price,font="arial 13 bold",bd=4,relief=SUNKEN).grid(row=0,column=1,padx=4,pady=4)

        m2_lbl=Label(F5,text="Total cow Milk Price",bg=bg_color,fg="White",font=("times new roman",13,"bold")).grid(row=1,column=0,padx=4,pady=4,sticky="w")
        m2_txt=Entry(F5,width=20,textvariable=self.cow_price,font="arial 13 bold",bd=4,relief=SUNKEN).grid(row=1,column=1,padx=4,pady=4)

        m3_lbl=Label(F5,text="Total Price",bg=bg_color,fg="White",font=("times new roman",13,"bold")).grid(row=2,column=0,padx=4,pady=4,sticky="w")
        m3_txt=Entry(F5,width=20,textvariable=self.totalboth,font="arial 13 bold",bd=4,relief=SUNKEN).grid(row=2,column=1,padx=4,pady=4)

        btn_F=Frame(F5,bd=7,relief=GROOVE)
        btn_F.place(x=420,width=880,height=105)

        total_btn=Button(btn_F,command=self.total,text="Total",bg="cadetblue",fg="white",bd=8,pady=17,width=11,font="arial 12 bold").grid(row=0,column=0,padx=5,pady=5)
        gbill_btn=Button(btn_F,command=self.bill_area,text="Genrate Bill",bg="cadetblue",fg="white",bd=8,pady=17,width=11,font="arial 12 bold").grid(row=0,column=1,padx=5,pady=5)
        save_btn=Button(btn_F,command=self.save,text="Save",bg="cadetblue",fg="white",bd=8,pady=17,width=11,font="arial 12 bold").grid(row=0,column=2,padx=5,pady=5)
        send_btn=Button(btn_F,command=self.btn,text="Send",bg="cadetblue",fg="white",bd=8,pady=17,width=11,font="arial 12 bold").grid(row=0,column=3,padx=5,pady=5)
        clear_btn=Button(btn_F,command=self.clear_data,text="Clear",bg="cadetblue",fg="white",bd=8,pady=17,width=11,font="arial 12 bold").grid(row=0,column=4,padx=5,pady=5)
        exit_btn=Button(btn_F,command=self.exit_app,text="Exit",bg="cadetblue",fg="white",bd=8,pady=17,width=11,font="arial 12 bold").grid(row=0,column=5,padx=5,pady=5)
        self.welcome_bill()
        #self.bill_n()
        
        
    def bill_n(self):
        file=openpyxl.load_workbook('Dairy_data.xlsx')
        sheet=file.active
        row=sheet.max_row

        self.max_row_value=sheet.cell(row=row,column=1).value
        #print(self.max_row_value)
        try:
            self.bill_no.set(self.max_row_value+1)
        except:
            self.bill_no.set("1")

        
        
    def total(self):
        self.bwp=self.bweight.get()*60
        self.total_baffelo=(
                            self.bwp
                            )
        self.baffelo_price.set(str(self.total_baffelo))

        self.bft=self.bfat.get()
        self.b_fat=(
                    self.bft
                    )
        self.baffelo_fat.set(str(self.b_fat))

        self.cwp=self.cweight.get()*40
        self.total_cow=(
                        self.cwp
                        )
        self.cow_price.set(str(self.total_cow))
        

        self.cft=self.cfat.get()

        self.c_fat=(
                    self.cft
                    )
        self.cow_fat.set(str(self.c_fat))

        self.Total_bill=float(self.total_baffelo+self.total_cow)

        self.totalbc=float(
                           self.total_baffelo+self.total_cow
                           )
        self.totalboth.set(str(self.totalbc))

        
    def welcome_bill(self):
        
        self.txtarea.delete('1.0',END)
        self.txtarea.insert(END,"\tWELCOME TO OM SAI DAIRY\n")
        self.txtarea.insert(END,f"\n Date : {self.Date.get()}")
        self.txtarea.insert(END,f"\n Bill Number : {self.bill_no.get()}")
        self.txtarea.insert(END,f"\n Customer Name : {self.c_name.get()}")
        self.txtarea.insert(END,f"\n Phone Number : {self.c_phon.get()}")
        self.txtarea.insert(END,f"\n********************************************")
        self.txtarea.insert(END,f"\n Milk\t\t   Liter\t\tPrice")
        self.txtarea.insert(END,f"\n********************************************")

        

    def bill_area(self):
        self.welcome_bill()
        self.bill_n()
        
        if self.c_name.get()=="" or self.c_phon.get()=="":
            messagebox.showerror("Error","Customer Details are Must")
        else:
            if self.bweight.get()!=0:
                self.txtarea.insert(END,f"\nBaffelo Weight\t\t   {self.bweight.get()}\t\t{self.bwp}")
            if self.bfat.get()!=0:
                self.txtarea.insert(END,f"\nBaffelo Fat\t\t   {self.bfat.get()}")
            if self.cweight.get()!=0:
                self.txtarea.insert(END,f"\nCow Weight\t\t   {self.cweight.get()}\t\t{self.cwp}")
            if self.cfat.get()!=0:
                self.txtarea.insert(END,f"\nCow Fat\t\t   {self.cfat.get()}")

            self.txtarea.insert(END,f"\n--------------------------------------------")
            self.txtarea.insert(END,f"\n \t\tTOTAL : {self.Total_bill}")
            self.txtarea.insert(END,f"\n--------------------------------------------")



            self.copy.set(str(self.Total_bill))
            self.cwp=self.cweight.get()
            self.copy1.set(str(self.cwp))
            self.bwp=self.bweight.get()
            self.copy2.set(str(self.bwp))
            self.cp=self.c_phon.get()
            self.copy3.set(str(self.cp))
             
    def btn(self):  
        self.account_sid = 'AC6fd618ad3cebf2306c0af28e0db06e2d'
        self.auth_token = 'd9bbf9c709cf5ba6c2f1d9c4526aac99'

        client = Client(self.account_sid, self.auth_token)
        
        self.message = client.messages.create(
                                               body= f"\n Date :- {self.Date.get()}\nBuffelo Weight :- {self.copy2.get()}\nrate :- 60\t\t\t\t\t\t\t\t\nCow Weight :- {self.copy1.get()}\nRate :- 40\t\t\t\t\t\t\t\nTotal :- {self.copy.get()}",
                                               from_='+12512775316',
                                               to={self.copy3.get()}
                                               )
        
        #print(message.sid)
        #self.kal.set(str(self.message))

    def clear_data(self):
        self.bweight.set(0)
        self.bfat.set(0)
        self.btemp.set(0)
        self.bsnf.set(0)
        self.bwater.set(0)
        self.brate.set(0)
        self.baverage.set(0)

        #cow
        self.cweight.set(0)
        self.cfat.set(0)
        self.ctemp.set(0)
        self.csnf.set(0)
        self.cwater.set(0)
        self.crate.set(0)
        self.caverage.set(0)

        #total
        self.baffelo_price.set("")
        self.cow_price.set("")
        self.totalboth.set("")
        self.baffelo_fat.set("")
        self.cow_fat.set("")
        self.total_bill.set("")
        
        
        #customer
        self.c_name.set("")
        self.c_phon.set("")
        #self.bill_no.set(0)
        self.search_bill.set("")
        self.kal.set("")
        self.copy.set("")
        self.copy1.set("")
        self.copy2.set("")
        self.copy3.set("")
        #self.Date.set("")
        self.welcome_bill()
        self.bill_n()
        

    def exit_app(self):
        op=messagebox.askyesno("Exit","Do you want Exit?")
        if op>0:
            self.root.destroy()

        
    file=pathlib.Path('Dairy_data.xlsx')
    if file.exists():
        pass
    else:
        file=Workbook()
        sheet =file.active
        sheet['A1']="Bill No."
        sheet['B1']="Customer Name"
        sheet['C1']="Phone No."
        sheet['D1']="Buffalo Milk Weight"
        sheet['E1']="Buffalo Fat"
        sheet['F1']="Cow Milk Weight"
        sheet['G1']="Cow Fat"
        sheet['H1']="Total Rs."
        sheet['I1']="Date"

        file.save('Dairy_data.xlsx')
        

    def save(self):
        self.R1=self.bill_no.get()
        try:    
            self.R2=self.c_name.get()
        except:
            messagebox.showerror("error","Enter the Name")
        self.R3=self.c_phon.get()
        self.R4=self.bweight.get()
        self.R5=self.bfat.get()
        self.R6=self.cweight.get()
        self.R7=self.cfat.get()
        self.R8=self.copy.get()
        self.R9=self.Date.get()

        if self.R2=="" or self.R3=="" or self.R8=="" or self.R4=="":
            messagebox.showerror("error","Data is Missing")
        else:
            file=openpyxl.load_workbook('Dairy_data.xlsx')
            sheet=file.active
            sheet.cell(column=1,row=sheet.max_row+1,value=self.R1)
            sheet.cell(column=2,row=sheet.max_row,value=self.R2)
            sheet.cell(column=3,row=sheet.max_row,value=self.R3)
            sheet.cell(column=4,row=sheet.max_row,value=self.R4)
            sheet.cell(column=5,row=sheet.max_row,value=self.R5)
            sheet.cell(column=6,row=sheet.max_row,value=self.R6)
            sheet.cell(column=7,row=sheet.max_row,value=self.R7)
            sheet.cell(column=8,row=sheet.max_row,value=self.R8)
            sheet.cell(column=9,row=sheet.max_row,value=self.R9)

            file.save(r'Dairy_data.xlsx')

            messagebox.showinfo("Info","Data has been saved")

            

    def fetch(self):

        text = self.search_bill.get()
        #self.bill_area()
        #self.clear_data()
        
        file=openpyxl.load_workbook("Dairy_data.xlsx")
        sheet=file.active

        for row in sheet.rows:
            if row[0].value == int(text):
                name=row[0]
                #print(str(name))
                self.reg_no_position=str(name)[14:-1]
                reg_number=str(name)[15:-1]

                #print(self.reg_no_position)
                #print(self.reg_number)
                           
        try:
            print(str(name))
        except:
            messagebox.showerror("invalid","invalid number")    

        self.x1=sheet.cell(row=int(reg_number),column=1).value
        self.x2=sheet.cell(row=int(reg_number),column=2).value
        self.x3=sheet.cell(row=int(reg_number),column=3).value
        self.x4=sheet.cell(row=int(reg_number),column=4).value
        self.x5=sheet.cell(row=int(reg_number),column=5).value
        self.x6=sheet.cell(row=int(reg_number),column=6).value
        self.x7=sheet.cell(row=int(reg_number),column=7).value
        self.x8=sheet.cell(row=int(reg_number),column=8).value
        self.x9=sheet.cell(row=int(reg_number),column=9).value
        
        #print(self.x1)
        #print(self.x2)
        #print(self.x3)
        #print(self.x4)
        #print(self.x5)
        #print(self.x6)
        #print(self.x7)
        #print(self.x8)
        #print(self.x9)

        #self.bill_area()
        #self.welcome_bill()
        #self.bill_no.set(self.x1)
        self.c_name.set(self.x2)
        self.c_phon.set(self.x3)
        self.bweight.set(self.x4)
        self.bfat.set(self.x5)
        self.cweight.set(self.x6)
        self.cfat.set(self.x7)
        self.totalboth.set(self.x8)
        #self.Date.set(self.x9)


root=Tk() 
obj = Bill_App(root)
root.mainloop()

 